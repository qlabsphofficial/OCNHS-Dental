from fastapi import FastAPI, Depends, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from database import SessionLocal, engine, Base
from pydantic import BaseModel
from typing import Optional, List
from models import Student, MedicalHistory, Appointment, Admin, OralHealthCondition, TemporaryTeeth, PermanentTeeth, DentalProcedure, ConditionTreatmentNeeds
from datetime import date, datetime, timedelta
from sqlalchemy import extract, cast, Date
from pytz import timezone
from fastapi.responses import FileResponse, RedirectResponse
from openpyxl import load_workbook
import smtplib
import os
from email.message import EmailMessage

email_address = "ocnhsdental.notification@gmail.com"
email_password = "xcwh bups wskc uasg"

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

Base.metadata.create_all(bind=engine)

def get_database():
    db = None

    try:
        db = SessionLocal()
    finally:
        yield db
        db.close()


class StudentModel(BaseModel):
    firstname: str
    middlename: Optional[str] = None
    lastname: str
    suffix: Optional[str] = None
    dateofbirth: Optional[date] = None
    gender: str
    age: Optional[int] = None
    birthplace: str
    contact_no: str
    address: str
    email_address: str
    password: str
    confirm_password: str

    parent_guardian_name: Optional[str] = None
    adviser_name: Optional[str] = None
    curriculum: Optional[str] = None
    grade_level: Optional[str] = None
    section: Optional[str] = None

    good_health: Optional[bool] = None
    under_medical_treatment: Optional[bool] = None
    condition_being_treated: Optional[str] = None
    serious_illness: Optional[bool] = None
    illness_or_operation: Optional[str] = None
    hospitalized: Optional[bool] = None
    hospitalization_details: Optional[str] = None
    taking_medication: Optional[bool] = None
    medication_details: Optional[str] = None
    use_tobacco: Optional[bool] = None
    use_alcohol_or_drugs: Optional[bool] = None
    pregnant_nursing_birth_control: Optional[bool] = None
    pregnant_nursing_birth_control_details: Optional[str] = None

    toothbrush: Optional[bool] = None
    brush_times_per_day: Optional[int] = None
    change_toothbrush_per_year: Optional[int] = None
    use_toothpaste: Optional[bool] = None
    dentist_visits_per_year: Optional[int] = None

    allergy: Optional[bool] = None
    # allergy_details: Optional[str] = None
    emphysema: Optional[bool] = None
    bleeding_problems: Optional[bool] = None
    blood_diseases: Optional[bool] = None
    head_injuries: Optional[bool] = None
    arthritis_rheumatism: Optional[bool] = None
    high_fever: Optional[bool] = None
    diabetes: Optional[bool] = None
    chest_pain: Optional[bool] = None
    stroke: Optional[bool] = None
    cancer_tumors: Optional[bool] = None
    anemia: Optional[bool] = None
    angina: Optional[bool] = None
    asthma: Optional[bool] = None
    high_blood_pressure: Optional[bool] = None
    low_blood_pressure: Optional[bool] = None
    aids_hiv_infection: Optional[bool] = None
    sexually_transmitted_disease: Optional[bool] = None
    stomach_troubles_ulcers: Optional[bool] = None
    fainting_seizure: Optional[bool] = None
    rapid_weight_loss_radiation_therapy: Optional[bool] = None
    joint_replacement_implant: Optional[bool] = None
    heart_surgery_heart_attack: Optional[bool] = None
    thyroid_problem: Optional[bool] = None
    heart_disease: Optional[bool] = None
    heart_murmur: Optional[bool] = None
    hepatitis_liver_disease: Optional[bool] = None
    rheumatic_seizure: Optional[bool] = None
    respiratory_problems: Optional[bool] = None
    hepatitis_jaundice: Optional[bool] = None
    tuberculosis: Optional[bool] = None
    swollen_ankles: Optional[bool] = None
    kidney_disease: Optional[bool] = None
    other_diseases: Optional[str] = None

    class Config:
        orm_mode = True


class LoginModel(BaseModel):
    email_address: str
    password: str


class AppointmentModel(BaseModel):
    student_id: int
    appointment_type: str
    appointment_datetime: datetime
    status: Optional[str] = 'PENDING'

    class Config:
        orm_mode = True
    
    
class StudentUpdateModel(BaseModel):
    student_id: int
    firstname: Optional[str] = None
    middlename: Optional[str] = None
    lastname: Optional[str] = None
    suffix: Optional[str] = None
    dateofbirth: Optional[date] = None
    gender: Optional[str] = None
    age: Optional[int] = None
    birthplace: Optional[str] = None
    contact_no: Optional[str] = None
    address: Optional[str] = None
    parent_guardian_name: Optional[str] = None
    adviser_name: Optional[str] = None
    curriculum: Optional[str] = None
    grade_level: Optional[str] = None
    section: Optional[str] = None

    class Config:
        orm_mode = True

    
class AdminRegisterModel(BaseModel):
    firstname: Optional[str] = None
    middlename: Optional[str] = None
    lastname: Optional[str] = None
    suffix: Optional[str] = None
    dateofbirth: Optional[date] = None
    gender: Optional[str] = None
    age: Optional[int] = None
    birthplace: Optional[str] = None
    contact_number: Optional[str] = None
    address: Optional[str] = None
    email: str
    password: str
    confirm_password: str

    class Config:
        orm_mode = True

    
class AdminLoginModel(BaseModel):
    email: str
    password: str


class AppointmentFilterRequest(BaseModel):
    year: int
    month: int


class UpdateMedicalRecordRequest(BaseModel):
    student_id: int

    good_health: Optional[bool] = None
    under_medical_treatment: Optional[bool] = None
    condition_being_treated: Optional[str] = None
    serious_illness: Optional[bool] = None
    illness_or_operation: Optional[str] = None
    hospitalized: Optional[bool] = None
    hospitalization_details: Optional[str] = None
    taking_medication: Optional[bool] = None
    medication_details: Optional[str] = None
    use_tobacco: Optional[bool] = None
    use_alcohol_or_drugs: Optional[bool] = None
    pregnant_nursing_birth_control: Optional[bool] = None
    pregnant_nursing_birth_control_details: Optional[str] = None

    toothbrush: Optional[bool] = None
    brush_times_per_day: Optional[int] = None
    change_toothbrush_per_year: Optional[int] = None
    use_toothpaste: Optional[bool] = None
    dentist_visits_per_year: Optional[int] = None

    allergy: Optional[bool] = None
    # allergy_details: Optional[str] = None
    emphysema: Optional[bool] = None
    bleeding_problems: Optional[bool] = None
    blood_diseases: Optional[bool] = None
    head_injuries: Optional[bool] = None
    arthritis_rheumatism: Optional[bool] = None
    high_fever: Optional[bool] = None
    diabetes: Optional[bool] = None
    chest_pain: Optional[bool] = None
    stroke: Optional[bool] = None
    cancer_tumors: Optional[bool] = None
    anemia: Optional[bool] = None
    angina: Optional[bool] = None
    asthma: Optional[bool] = None
    high_blood_pressure: Optional[bool] = None
    low_blood_pressure: Optional[bool] = None
    aids_hiv_infection: Optional[bool] = None
    sexually_transmitted_disease: Optional[bool] = None
    stomach_troubles_ulcers: Optional[bool] = None
    fainting_seizure: Optional[bool] = None
    rapid_weight_loss_radiation_therapy: Optional[bool] = None
    joint_replacement_implant: Optional[bool] = None
    heart_surgery_heart_attack: Optional[bool] = None
    thyroid_problem: Optional[bool] = None
    heart_disease: Optional[bool] = None
    heart_murmur: Optional[bool] = None
    hepatitis_liver_disease: Optional[bool] = None
    rheumatic_seizure: Optional[bool] = None
    respiratory_problems: Optional[bool] = None
    hepatitis_jaundice: Optional[bool] = None
    tuberculosis: Optional[bool] = None
    swollen_ankles: Optional[bool] = None
    kidney_disease: Optional[bool] = None
    other_diseases: Optional[str] = None

    class Config:
        orm_mode = True
    
    
class ApproveAppointmentRequest(BaseModel):
    appointment_id: int
    
    
class CancelAppointmentRequest(BaseModel):
    appointment_id: int
    
    
class RescheduleAppointmentRequest(BaseModel):
    appointment_id: int
    date: datetime


class StudentFilterRequest(BaseModel):
    is_archive: bool
    year: str | None
    curriculum: str | None
    grade_level: str | None
    section: str | None


class ArchiveUpdateRequest(BaseModel):
    student_id: int
    is_archive: bool
    

class ActiveUpdateRequest(BaseModel):
    student_id: int
    is_active: bool


class ForgotPasswordRequest(BaseModel):
    email: str


class StudentRequest(BaseModel):
    student_id: int
    
    
class UpdateStudentMedicalHistory(BaseModel):
    student_id: int

    good_health: Optional[bool] = None
    under_medical_treatment: Optional[bool] = None
    condition_being_treated: Optional[str] = None
    serious_illness: Optional[bool] = None
    illness_or_operation: Optional[str] = None
    hospitalized: Optional[bool] = None
    hospitalization_details: Optional[str] = None
    taking_medication: Optional[bool] = None
    medication_details: Optional[str] = None
    use_tobacco: Optional[bool] = None
    use_alcohol_or_drugs: Optional[bool] = None
    pregnant_nursing_birth_control: Optional[bool] = None
    pregnant_nursing_birth_control_details: Optional[str] = None

    toothbrush: Optional[bool] = None
    brush_times_per_day: Optional[int] = None
    change_toothbrush_per_year: Optional[int] = None
    use_toothpaste: Optional[bool] = None
    dentist_visits_per_year: Optional[int] = None

    class Config:
        orm_mode = True
        
    
class ResetPasswordRequest(BaseModel):
    new_password: str
    confirm_new_password: str
    
    
class DeleteStudentRecord(BaseModel):
    student_id: int


class TemporaryTeethSchema(BaseModel):
    student_id: int
    grade_level: str
    no_t_decayed: str
    no_t_filled: str
    total_dft: str
    
    
class PermanentTeethSchema(BaseModel):
    student_id: int
    grade_level: str
    no_t_decayed: str
    no_t_missing: str
    no_t_filled: str
    total_d_m_f_t: str
    total_sound_teeth: str
    
    
class OralHealthConditionSchema(BaseModel):
    student_id: int
    grade_level: str
    gingivitis: str
    periodontal_disease: str
    malocclusion: str
    supernumerary_teeth: str
    retained_deciduous_teeth: str
    decubital_ulcer: str
    calculus: str
    cleft_lip_palate: str
    root_fragment: str
    fluorosis: str
    others: str
    
    
class OralHealthConditionSchema(BaseModel):
    student_id: int
    grade_level: str
    gingivitis: str
    periodontal_disease: str
    malocclusion: str
    supernumerary_teeth: str
    retained_deciduous_teeth: str
    decubital_ulcer: str
    calculus: str
    cleft_lip_palate: str
    root_fragment: str
    fluorosis: str
    others: str
    
    
class DentalProcedureSchema(BaseModel):
    student_id: int
    grade_level: str
    date: str
    examination: str
    sealant_gi: str
    gum_treatment: str
    permanent_filling: str
    art: str
    extraction: str
    oral_prophylaxis: str
    referral: str
    other_oral_treatment: str
    examined_by: str
    
    
class ConditionTreatmentNeedsSchema(BaseModel):
    student_id: int
    layer_no: int
    cell_no: int
    value: str
    
    
class ConditionTreatmentNeedsArraySchema(BaseModel):
    student_id: int
    layer_no: int
    cell_no: int
    values: List[str]
    
    
@app.post("/register")
async def register(student: StudentModel, db: Session = Depends(get_database)):
    if student.password != student.confirm_password:
        raise HTTPException(status_code=400, detail="Passwords do not match")

    existing_student = db.query(Student).filter(Student.email_address == student.email_address).first()
    if existing_student:
        raise HTTPException(status_code=400, detail="Email already registered")

    new_student = Student(
        firstname=student.firstname,
        middlename=student.middlename,
        lastname=student.lastname,
        suffix=student.suffix,
        dateofbirth=student.dateofbirth,
        gender=student.gender,
        age=student.age,
        birthplace=student.birthplace,
        contact_no=student.contact_no,
        address=student.address,
        email_address=student.email_address,
        password=student.password,

        parent_guardian_name=student.parent_guardian_name,
        adviser_name=student.adviser_name,
        curriculum=student.curriculum,
        grade_level=student.grade_level,
        section=student.section,
    )

    db.add(new_student)
    db.commit()
    db.refresh(new_student)

    new_medical_history = MedicalHistory(
        student_id=new_student.id,
        good_health=student.good_health,
        under_medical_treatment=student.under_medical_treatment,
        condition_being_treated=student.condition_being_treated,
        serious_illness=student.serious_illness,
        illness_or_operation=student.illness_or_operation,
        hospitalized=student.hospitalized,
        hospitalization_details=student.hospitalization_details,
        taking_medication=student.taking_medication,
        medication_details=student.medication_details,
        use_tobacco=student.use_tobacco,
        use_alcohol_or_drugs=student.use_alcohol_or_drugs,
        pregnant_nursing_birth_control=student.pregnant_nursing_birth_control,
        pregnant_nursing_birth_control_details=student.pregnant_nursing_birth_control_details,
        toothbrush=student.toothbrush,
        brush_times_per_day=student.brush_times_per_day,
        change_toothbrush_per_year=student.change_toothbrush_per_year,
        use_toothpaste=student.use_toothpaste,
        dentist_visits_per_year=student.dentist_visits_per_year,
        allergy=student.allergy,
        # allergy_details=student.allergy_details,
        emphysema=student.emphysema,
        bleeding_problems=student.bleeding_problems,
        blood_diseases=student.blood_diseases,
        head_injuries=student.head_injuries,
        arthritis_rheumatism=student.arthritis_rheumatism,
        high_fever=student.high_fever,
        diabetes=student.diabetes,
        chest_pain=student.chest_pain,
        stroke=student.stroke,
        cancer_tumors=student.cancer_tumors,
        anemia=student.anemia,
        angina=student.angina,
        asthma=student.asthma,
        high_blood_pressure=student.high_blood_pressure,
        low_blood_pressure=student.low_blood_pressure,
        aids_hiv_infection=student.aids_hiv_infection,
        sexually_transmitted_disease=student.sexually_transmitted_disease,
        stomach_troubles_ulcers=student.stomach_troubles_ulcers,
        fainting_seizure=student.fainting_seizure,
        rapid_weight_loss_radiation_therapy=student.rapid_weight_loss_radiation_therapy,
        joint_replacement_implant=student.joint_replacement_implant,
        heart_surgery_heart_attack=student.heart_surgery_heart_attack,
        thyroid_problem=student.thyroid_problem,
        heart_disease=student.heart_disease,
        heart_murmur=student.heart_murmur,
        hepatitis_liver_disease=student.hepatitis_liver_disease,
        rheumatic_seizure=student.rheumatic_seizure,
        respiratory_problems=student.respiratory_problems,
        hepatitis_jaundice=student.hepatitis_jaundice,
        tuberculosis=student.tuberculosis,
        swollen_ankles=student.swollen_ankles,
        kidney_disease=student.kidney_disease,
        other_diseases=student.other_diseases
    )
    db.add(new_medical_history)
    db.commit()
    db.refresh(new_medical_history)

    return {"message": "Registration successful", "status_code": 200, "new_student_id": new_student.id}

    
@app.post("/login")
async def login(login_data: LoginModel, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.email_address == login_data.email_address).first()

    if not student or student.password != login_data.password:
        raise HTTPException(status_code=400, detail="Invalid email or password")
    
    medical_history = db.query(MedicalHistory).filter(MedicalHistory.student_id == student.id).first()

    if medical_history:
        return {
            "message": "Login successful",
            "student_data": student,
            "medical_history": medical_history
        }
    else:
        return {
            "message": "Login successful",
            "student_data": student,
            "medical_history": "Does not exist!"
        }


@app.post("/create_appointment")
async def create_appointment(appointment_data: AppointmentModel, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.id == appointment_data.student_id).first()

    if not student:
        raise HTTPException(status_code=400, detail="Student not found")
    
    new_appointment = Appointment(
        student_id=appointment_data.student_id,
        appointment_type=appointment_data.appointment_type,
        appointment_datetime=appointment_data.appointment_datetime,
        status=appointment_data.status
    )

    db.add(new_appointment)
    db.commit()
    db.refresh(new_appointment)

    return {
        "message": "Appointment created successfully",
    }


@app.get("/get_student_appointments")
async def get_student_appointments(student_id: int, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.id == student_id).first()
    
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    appointments = db.query(Appointment).filter(Appointment.student_id == student_id).all()
    
    if not appointments:
        return {"message": "No appointments found for this student"}
    
    return {"appointments": appointments}


@app.put("update_student")
async def update_student(student_data: StudentUpdateModel, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.id == student_data.student_id).first()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    student.firstname = student_data.firstname
    student.middlename = student_data.middlename
    student.lastname = student_data.lastname
    student.suffix = student_data.suffix
    student.dateofbirth = student_data.dateofbirth
    student.gender = student_data.gender
    student.age = student_data.age
    student.birthplace = student_data.birthplace
    student.contact_no = student_data.contact_no
    student.address = student_data.address
    student.parent_guardian_name = student_data.parent_guardian_name
    student.adviser_name = student_data.adviser_name
    student.curriculum = student_data.curriculum
    student.grade_level = student_data.grade_level
    student.section = student_data.section

    db.commit()
    db.refresh(student)

    return {"message": "Student updated successfully", "student_data": student}


@app.get("/get_medical_history")
async def get_medical_history(student_id: int, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.id == student_id).first()
    
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    medical_history = db.query(MedicalHistory).filter(MedicalHistory.student_id == student_id).all()
    
    if not medical_history:
        return {"message": "No medical history found for this student"}
    
    return {"medical_history": medical_history}


@app.get("/get_all_medical_history")
async def get_all_medical_history(db: Session = Depends(get_database)):
    medical_history = db.query(MedicalHistory).all()
    
    if not medical_history:
        return {"message": "No medical history found for this student"}
    
    return {"medical_history": medical_history}


@app.post("/register_admin")
async def register_admin(admin_data: AdminRegisterModel, db: Session = Depends(get_database)):
    try:
        if admin_data.password != admin_data.confirm_password:
            raise HTTPException(status_code=400, detail="Passwords do not match")
        
        existing_admin = db.query(Admin).filter(Admin.email == admin_data.email).first()
        if existing_admin:
            raise HTTPException(status_code=400, detail="Email already registered")

        new_admin = Admin(
            firstname=admin_data.firstname,
            middlename=admin_data.middlename,
            lastname=admin_data.lastname,
            suffix=admin_data.suffix,
            dateofbirth=admin_data.dateofbirth,
            gender=admin_data.gender,
            age=admin_data.age,
            birthplace=admin_data.birthplace,
            contact_number=admin_data.contact_number,
            address=admin_data.address,
            email=admin_data.email,
            password=admin_data.password
        )

        db.add(new_admin)
        db.commit()
        db.refresh(new_admin)

        return {"message": "Admin registration successful", "status_code": 200}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Registration failed: {str(e)}")
    

@app.post("/login_admin")
async def login_admin(login_data: AdminLoginModel, db: Session = Depends(get_database)):
    admin = db.query(Admin).filter(Admin.email == login_data.email).first()

    if not admin or admin.password != login_data.password:
        raise HTTPException(status_code=400, detail="Invalid email or password")
    
    return {
        "message": "Login successful",
        "admin_data": admin
    }


@app.get("/get_pending_appointments")
async def get_pending_appointments(db: Session = Depends(get_database)):
    results = (
        db.query(Appointment, Student)
        .join(Student, Student.id == Appointment.student_id)  # INNER JOIN
        .filter(Appointment.status == 'PENDING')
        .all()
    )

    # Format results nicely
    formatted_results = []
    for appointment, student in results:
        formatted_results.append({
        # Student fields
        "student_id": student.id if student else None,
        "firstname": student.firstname if student else None,
        "middlename": student.middlename if student else None,
        "lastname": student.lastname if student else None,
        "suffix": student.suffix if student else None,
        "dateofbirth": student.dateofbirth if student else None,
        "gender": student.gender if student else None,
        "age": student.age if student else None,
        "birthplace": student.birthplace if student else None,
        "contact_no": student.contact_no if student else None,
        "address": student.address if student else None,
        "email_address": student.email_address if student else None,
        "password": student.password if student else None,
        "parent_guardian_name": student.parent_guardian_name if student else None,
        "adviser_name": student.adviser_name if student else None,
        "curriculum": student.curriculum if student else None,
        "grade_level": student.grade_level if student else None,
        "section": student.section if student else None,

        # Appointment fields
        "appointment_id": appointment.id,
        "appointment_type": appointment.appointment_type,
        "appointment_datetime": appointment.appointment_datetime,
        "status": appointment.status,
    })

    return {"pending_appointments": formatted_results}


@app.get("/appointments_by_month") 
async def get_appointments_by_month(filter: AppointmentFilterRequest, db: Session = Depends(get_database)):
    year = filter.year
    month = filter.month
    
    # Query appointments with the extracted year and month
    results = (
        db.query(Appointment, Student)
        .join(Student, Student.id == Appointment.student_id)  # INNER JOIN to ensure only matching students
        .filter(
            extract('year', Appointment.appointment_datetime) == year,
            extract('month', Appointment.appointment_datetime) == month
        )
        .all()
    )
    
    if not results:
        raise HTTPException(status_code=404, detail="No appointments found for the given month and year")
    
    formatted_results = []
    for appointment, student in results:
        formatted_results.append({
            # Student fields
        "student_id": student.id if student else None,
        "firstname": student.firstname if student else None,
        "middlename": student.middlename if student else None,
        "lastname": student.lastname if student else None,
        "suffix": student.suffix if student else None,
        "dateofbirth": student.dateofbirth if student else None,
        "gender": student.gender if student else None,
        "age": student.age if student else None,
        "birthplace": student.birthplace if student else None,
        "contact_no": student.contact_no if student else None,
        "address": student.address if student else None,
        "email_address": student.email_address if student else None,
        "password": student.password if student else None,
        "parent_guardian_name": student.parent_guardian_name if student else None,
        "adviser_name": student.adviser_name if student else None,
        "curriculum": student.curriculum if student else None,
        "grade_level": student.grade_level if student else None,
        "section": student.section if student else None,

        # Appointment fields
        "appointment_id": appointment.id,
        "appointment_type": appointment.appointment_type,
        "appointment_datetime": appointment.appointment_datetime,
        "status": appointment.status,
        })

    return {"appointments": formatted_results}


@app.get("/get_today_appointments")
async def get_today_appointments(db: Session = Depends(get_database)):
    manila_tz = timezone("Asia/Manila")
    now = datetime.now(manila_tz)

    start_of_day = datetime(now.year, now.month, now.day, 0, 0, 0, tzinfo=manila_tz)
    end_of_day = start_of_day + timedelta(days=1)

    appointments = (
        db.query(Appointment, Student)
        .join(Student, Student.id == Appointment.student_id)
        .filter(
            Appointment.appointment_datetime >= start_of_day,
            Appointment.appointment_datetime < end_of_day,
            Appointment.status == 'APPROVED'
        )
        .all()
    )
    
    if not appointments:
        return {"appointments": []}
    
    formatted_appointments = [
        {
            "appointment_id": appointment.id,
            "student_id": appointment.student_id,
            "appointment_type": appointment.appointment_type,
            "appointment_datetime": appointment.appointment_datetime,
            "status": appointment.status,
            "student_info": student,
        }
        for appointment, student in appointments
    ]
    
    return {"appointments": formatted_appointments}


@app.get("/get_all_appointments")
async def get_all_appointments(db: Session = Depends(get_database)):
    appointments = db.query(Appointment, Student).join(Student, Student.id == Appointment.student_id).all()
    
    if not appointments:
        raise HTTPException(status_code=404, detail="No appointments found")
    
    formatted_appointments = []
    for appointment, student in appointments:
        formatted_appointments.append({
            "appointment_id": appointment.id,
            "student_id": appointment.student_id,
            "appointment_type": appointment.appointment_type,
            "appointment_datetime": appointment.appointment_datetime,
            "status": appointment.status,
            'student_info': student
        })
    
    return {"appointments": formatted_appointments, "count": len(formatted_appointments)}


@app.post("/update_medical_record")
async def update_medical_record(medical_history_data: UpdateMedicalRecordRequest, db: Session = Depends(get_database)):
    # Check if the student exists
    student = db.query(Student).filter(Student.id == medical_history_data.student_id).first()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    medical_history = db.query(MedicalHistory).filter(MedicalHistory.student_id == medical_history_data.student_id).first()
    
    if not medical_history:
        raise HTTPException(status_code=404, detail="Medical record not found for the student")

    else:
        medical_history.good_health = medical_history_data.good_health
        medical_history.under_medical_treatment = medical_history_data.under_medical_treatment
        medical_history.condition_being_treated = medical_history_data.condition_being_treated
        medical_history.serious_illness = medical_history_data.serious_illness
        medical_history.illness_or_operation = medical_history_data.illness_or_operation
        medical_history.hospitalized = medical_history_data.hospitalized
        medical_history.hospitalization_details = medical_history_data.hospitalization_details
        medical_history.taking_medication = medical_history_data.taking_medication
        medical_history.medication_details = medical_history_data.medication_details
        medical_history.use_tobacco = medical_history_data.use_tobacco
        medical_history.use_alcohol_or_drugs = medical_history_data.use_alcohol_or_drugs
        medical_history.pregnant_nursing_birth_control = medical_history_data.pregnant_nursing_birth_control
        medical_history.pregnant_nursing_birth_control_details = medical_history_data.pregnant_nursing_birth_control_details
        medical_history.toothbrush = medical_history_data.toothbrush
        medical_history.brush_times_per_day = medical_history_data.brush_times_per_day
        medical_history.change_toothbrush_per_year = medical_history_data.change_toothbrush_per_year
        medical_history.use_toothpaste = medical_history_data.use_toothpaste
        medical_history.dentist_visits_per_year = medical_history_data.dentist_visits_per_year
        medical_history.allergy = medical_history_data.allergy
        # medical_history.allergy_details = medical_history_data.allergy_details
        medical_history.emphysema = medical_history_data.emphysema
        medical_history.bleeding_problems = medical_history_data.bleeding_problems
        medical_history.blood_diseases = medical_history_data.blood_diseases
        medical_history.head_injuries = medical_history_data.head_injuries
        medical_history.arthritis_rheumatism = medical_history_data.arthritis_rheumatism
        medical_history.high_fever = medical_history_data.high_fever
        medical_history.diabetes = medical_history_data.diabetes
        medical_history.chest_pain = medical_history_data.chest_pain
        medical_history.stroke = medical_history_data.stroke
        medical_history.cancer_tumors = medical_history_data.cancer_tumors
        medical_history.anemia = medical_history_data.anemia
        medical_history.angina = medical_history_data.angina
        medical_history.asthma = medical_history_data.asthma
        medical_history.high_blood_pressure = medical_history_data.high_blood_pressure
        medical_history.low_blood_pressure = medical_history_data.low_blood_pressure
        medical_history.aids_hiv_infection = medical_history_data.aids_hiv_infection
        medical_history.sexually_transmitted_disease = medical_history_data.sexually_transmitted_disease
        medical_history.stomach_troubles_ulcers = medical_history_data.stomach_troubles_ulcers
        medical_history.fainting_seizure = medical_history_data.fainting_seizure
        medical_history.rapid_weight_loss_radiation_therapy = medical_history_data.rapid_weight_loss_radiation_therapy
        medical_history.joint_replacement_implant = medical_history_data.joint_replacement_implant
        medical_history.heart_surgery_heart_attack = medical_history_data.heart_surgery_heart_attack
        medical_history.thyroid_problem = medical_history_data.thyroid_problem
        medical_history.heart_disease = medical_history_data.heart_disease
        medical_history.heart_murmur = medical_history_data.heart_murmur
        medical_history.hepatitis_liver_disease = medical_history_data.hepatitis_liver_disease
        medical_history.rheumatic_seizure = medical_history_data.rheumatic_seizure
        medical_history.respiratory_problems = medical_history_data.respiratory_problems
        medical_history.hepatitis_jaundice = medical_history_data.hepatitis_jaundice
        medical_history.tuberculosis = medical_history_data.tuberculosis
        medical_history.swollen_ankles = medical_history_data.swollen_ankles
        medical_history.kidney_disease = medical_history_data.kidney_disease
        medical_history.other_diseases = medical_history_data.other_diseases
        
        db.commit()
        db.refresh(medical_history)

        return {"message": "Medical history updated successfully", "medical_history": medical_history}
    
    
@app.put("/approve_appointment")
async def approve_appointment(request: ApproveAppointmentRequest, db: Session = Depends(get_database)):
    appointment = db.query(Appointment).filter(Appointment.id == request.appointment_id).first()

    if not appointment:
        raise HTTPException(status_code=404, detail="Appointment not found")

    appointment.status = "APPROVED"
    db.commit()
    db.refresh(appointment)
    
    student = db.query(Student).filter(Student.id == appointment.student_id).first()
    
    formatted_appointment_datetime = appointment.appointment_datetime.strftime("%A, %B %d, %Y at %I:%M %p")
    
    msg = EmailMessage()
    msg['Subject'] = f"Your OCNHS Dental Appointment – {formatted_appointment_datetime}"
    msg['From'] = email_address
    msg['To'] = student.email_address
    msg.set_content(
    f"""\
    Hi {student.firstname} {student.lastname},

    Your appointment for {formatted_appointment_datetime} has been approved.

    This is an automated message—please do not reply.

    Thanks,
    OCNHS Dental Team
    """,

    )
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(email_address, email_password)
        smtp.send_message(msg)
        

    return { "message": "Appointment approved successfully" }


@app.put("/cancel_appointment")
async def cancel_appointment(request: CancelAppointmentRequest, db: Session = Depends(get_database)):
    appointment = db.query(Appointment).filter(Appointment.id == request.appointment_id).first()

    if not appointment:
        raise HTTPException(status_code=404, detail="Appointment not found")

    appointment.status = "CANCELLED"
    db.commit()
    db.refresh(appointment)

    student = db.query(Student).filter(Student.id == appointment.student_id).first()
    
    formatted_appointment_datetime = appointment.appointment_datetime.strftime("%A, %B %d, %Y at %I:%M %p")
    
    msg = EmailMessage()
    msg['Subject'] = f"Your OCNHS Dental Appointment – {formatted_appointment_datetime}"
    msg['From'] = email_address
    msg['To'] = student.email_address
    msg.set_content(
    f"""\
    Hi {student.firstname} {student.lastname},

    Your appointment for {formatted_appointment_datetime} has been canceled.

    This is an automated message—please do not reply.

    Thanks,
    OCNHS Dental Team
    """,

    )
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(email_address, email_password)
        smtp.send_message(msg)

    return { "message": "Appointment cancelled successfully" }


@app.put("/reschedule_appointment")
async def reschedule_appointment(request: RescheduleAppointmentRequest, db: Session = Depends(get_database)):
    appointment = db.query(Appointment).filter(Appointment.id == request.appointment_id).first()
    formatted_appointment_datetime = appointment.appointment_datetime.strftime("%A, %B %d, %Y at %I:%M %p")
    
    if not appointment:
        raise HTTPException(status_code=404, detail="Appointment not found")

    appointment.status = "RESCHEDULED"
    appointment.appointment_datetime = request.date
    db.commit()
    db.refresh(appointment)

    student = db.query(Student).filter(Student.id == appointment.student_id).first()
    
    formatted_new_appointment_datetime = request.date.strftime("%A, %B %d, %Y at %I:%M %p")

    msg = EmailMessage()
    msg['Subject'] = f"Your OCNHS Dental Appointment – {formatted_appointment_datetime}"
    msg['From'] = email_address
    msg['To'] = student.email_address
    msg.set_content(
    f"""\
    Hi {student.firstname} {student.lastname},

    Your appointment has been rescheduled to {formatted_new_appointment_datetime}.

    This is an automated message—please do not reply.

    Thanks,
    OCNHS Dental Team
    """,

    )
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(email_address, email_password)
        smtp.send_message(msg)

    return { "message": "Appointment rescheduled successfully" }


@app.post("/get_students")
async def get_students(filters: StudentFilterRequest, db: Session = Depends(get_database)):
    if filters.is_archive:
        query = db.query(Student).filter(Student.is_archive == filters.is_archive)

        if filters.year:
            query = query.filter(extract('year', Student.date_archived) == filters.year)
        if filters.curriculum:
            query = query.filter(Student.curriculum == filters.curriculum)

        students = query.all()
    else:
        query = db.query(Student).filter(Student.is_archive == filters.is_archive)

        if filters.curriculum:
            query = query.filter(Student.curriculum == filters.curriculum)
        if filters.grade_level:
            query = query.filter(Student.grade_level == filters.grade_level)
        if filters.section:
            query = query.filter(Student.section == filters.section)

        students = query.all()

    return {"students": students}


@app.post("/get_student_medical_history")
async def get_student_medical_history(data: StudentRequest, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.id == data.student_id).first()
    
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    medical_history = db.query(MedicalHistory).filter(MedicalHistory.student_id == data.student_id).first()

    if not medical_history:
        return {
            "student": {
                "id": student.id,
                "firstname": student.firstname,
                "middlename": student.middlename,
                "lastname": student.lastname,
                "suffix": student.suffix,
                "dateofbirth": student.dateofbirth,
                "gender": student.gender,
                "age": student.age,
                "birthplace": student.birthplace,
                "contact_no": student.contact_no,
                "address": student.address,
                "email_address": student.email_address,
                "parent_guardian_name": student.parent_guardian_name,
                "adviser_name": student.adviser_name,
                "curriculum": student.curriculum,
                "grade_level": student.grade_level,
                "section": student.section,
                "is_archive": student.is_archive,
                "is_active": student.is_active,
            },
            "medical_history": None
        }

    return {
        "student": {
            "id": student.id,
            "firstname": student.firstname,
            "middlename": student.middlename,
            "lastname": student.lastname,
            "suffix": student.suffix,
            "dateofbirth": student.dateofbirth,
            "gender": student.gender,
            "age": student.age,
            "birthplace": student.birthplace,
            "contact_no": student.contact_no,
            "address": student.address,
            "email_address": student.email_address,
            "parent_guardian_name": student.parent_guardian_name,
            "adviser_name": student.adviser_name,
            "curriculum": student.curriculum,
            "grade_level": student.grade_level,
            "section": student.section,
            "is_archive": student.is_archive,
            "is_active": student.is_active,
        },
        "medical_history": {
            "id": medical_history.id,
            "student_id": medical_history.student_id,
            "good_health": medical_history.good_health,
            "under_medical_treatment": medical_history.under_medical_treatment,
            "condition_being_treated": medical_history.condition_being_treated,
            "serious_illness": medical_history.serious_illness,
            "illness_or_operation": medical_history.illness_or_operation,
            "hospitalized": medical_history.hospitalized,
            "hospitalization_details": medical_history.hospitalization_details,
            "taking_medication": medical_history.taking_medication,
            "medication_details": medical_history.medication_details,
            "use_tobacco": medical_history.use_tobacco,
            "use_alcohol_or_drugs": medical_history.use_alcohol_or_drugs,
            "pregnant_nursing_birth_control": medical_history.pregnant_nursing_birth_control,
            "pregnant_nursing_birth_control_details": medical_history.pregnant_nursing_birth_control_details,
            "toothbrush": medical_history.toothbrush,
            "brush_times_per_day": medical_history.brush_times_per_day,
            "change_toothbrush_per_year": medical_history.change_toothbrush_per_year,
            "use_toothpaste": medical_history.use_toothpaste,
            "dentist_visits_per_year": medical_history.dentist_visits_per_year,
        }
    }
    
    
@app.post("/update_student_medical_history")
async def update_student_medical_history(medical_history_data: UpdateStudentMedicalHistory, db: Session = Depends(get_database)):
    # Check if the student exists
    student = db.query(Student).filter(Student.id == medical_history_data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Check for existing medical history
    medical_history = db.query(MedicalHistory).filter(MedicalHistory.student_id == medical_history_data.student_id).first()

    if not medical_history:
        medical_history = MedicalHistory(
            student_id=medical_history_data.student_id,
            good_health=medical_history_data.good_health,
            under_medical_treatment=medical_history_data.under_medical_treatment,
            condition_being_treated=medical_history_data.condition_being_treated,
            serious_illness=medical_history_data.serious_illness,
            illness_or_operation=medical_history_data.illness_or_operation,
            hospitalized=medical_history_data.hospitalized,
            hospitalization_details=medical_history_data.hospitalization_details,
            taking_medication=medical_history_data.taking_medication,
            medication_details=medical_history_data.medication_details,
            use_tobacco=medical_history_data.use_tobacco,
            use_alcohol_or_drugs=medical_history_data.use_alcohol_or_drugs,
            pregnant_nursing_birth_control=medical_history_data.pregnant_nursing_birth_control,
            pregnant_nursing_birth_control_details=medical_history_data.pregnant_nursing_birth_control_details,
            toothbrush=medical_history_data.toothbrush,
            brush_times_per_day=medical_history_data.brush_times_per_day,
            change_toothbrush_per_year=medical_history_data.change_toothbrush_per_year,
            use_toothpaste=medical_history_data.use_toothpaste,
            dentist_visits_per_year=medical_history_data.dentist_visits_per_year
        )
        db.add(medical_history)
        message = "Medical history created successfully"
    else:
        medical_history.good_health = medical_history_data.good_health
        medical_history.under_medical_treatment = medical_history_data.under_medical_treatment
        medical_history.condition_being_treated = medical_history_data.condition_being_treated
        medical_history.serious_illness = medical_history_data.serious_illness
        medical_history.illness_or_operation = medical_history_data.illness_or_operation
        medical_history.hospitalized = medical_history_data.hospitalized
        medical_history.hospitalization_details = medical_history_data.hospitalization_details
        medical_history.taking_medication = medical_history_data.taking_medication
        medical_history.medication_details = medical_history_data.medication_details
        medical_history.use_tobacco = medical_history_data.use_tobacco
        medical_history.use_alcohol_or_drugs = medical_history_data.use_alcohol_or_drugs
        medical_history.pregnant_nursing_birth_control = medical_history_data.pregnant_nursing_birth_control
        medical_history.pregnant_nursing_birth_control_details = medical_history_data.pregnant_nursing_birth_control_details
        medical_history.toothbrush = medical_history_data.toothbrush
        medical_history.brush_times_per_day = medical_history_data.brush_times_per_day
        medical_history.change_toothbrush_per_year = medical_history_data.change_toothbrush_per_year
        medical_history.use_toothpaste = medical_history_data.use_toothpaste
        medical_history.dentist_visits_per_year = medical_history_data.dentist_visits_per_year
        message = "Medical history updated successfully"

    db.commit()
    db.refresh(medical_history)

    return {"message": message, "medical_history": medical_history, "status_code": 200 }
    
    
@app.post("/update_medical_record")
async def update_medical_record(medical_history_data: UpdateMedicalRecordRequest, db: Session = Depends(get_database)):
    # Check if the student exists
    student = db.query(Student).filter(Student.id == medical_history_data.student_id).first()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    medical_history = db.query(MedicalHistory).filter(MedicalHistory.student_id == medical_history_data.student_id).first()
    
    if not medical_history:
        raise HTTPException(status_code=404, detail="Medical record not found for the student")

    else:
        medical_history.good_health = medical_history_data.good_health
        medical_history.under_medical_treatment = medical_history_data.under_medical_treatment
        medical_history.condition_being_treated = medical_history_data.condition_being_treated
        medical_history.serious_illness = medical_history_data.serious_illness
        medical_history.illness_or_operation = medical_history_data.illness_or_operation
        medical_history.hospitalized = medical_history_data.hospitalized
        medical_history.hospitalization_details = medical_history_data.hospitalization_details
        medical_history.taking_medication = medical_history_data.taking_medication
        medical_history.medication_details = medical_history_data.medication_details
        medical_history.use_tobacco = medical_history_data.use_tobacco
        medical_history.use_alcohol_or_drugs = medical_history_data.use_alcohol_or_drugs
        medical_history.pregnant_nursing_birth_control = medical_history_data.pregnant_nursing_birth_control
        medical_history.pregnant_nursing_birth_control_details = medical_history_data.pregnant_nursing_birth_control_details
        medical_history.toothbrush = medical_history_data.toothbrush
        medical_history.brush_times_per_day = medical_history_data.brush_times_per_day
        medical_history.change_toothbrush_per_year = medical_history_data.change_toothbrush_per_year
        medical_history.use_toothpaste = medical_history_data.use_toothpaste
        medical_history.dentist_visits_per_year = medical_history_data.dentist_visits_per_year
        medical_history.allergy = medical_history_data.allergy
        # medical_history.allergy_details = medical_history_data.allergy_details
        medical_history.emphysema = medical_history_data.emphysema
        medical_history.bleeding_problems = medical_history_data.bleeding_problems
        medical_history.blood_diseases = medical_history_data.blood_diseases
        medical_history.head_injuries = medical_history_data.head_injuries
        medical_history.arthritis_rheumatism = medical_history_data.arthritis_rheumatism
        medical_history.high_fever = medical_history_data.high_fever
        medical_history.diabetes = medical_history_data.diabetes
        medical_history.chest_pain = medical_history_data.chest_pain
        medical_history.stroke = medical_history_data.stroke
        medical_history.cancer_tumors = medical_history_data.cancer_tumors
        medical_history.anemia = medical_history_data.anemia
        medical_history.angina = medical_history_data.angina
        medical_history.asthma = medical_history_data.asthma
        medical_history.high_blood_pressure = medical_history_data.high_blood_pressure
        medical_history.low_blood_pressure = medical_history_data.low_blood_pressure
        medical_history.aids_hiv_infection = medical_history_data.aids_hiv_infection
        medical_history.sexually_transmitted_disease = medical_history_data.sexually_transmitted_disease
        medical_history.stomach_troubles_ulcers = medical_history_data.stomach_troubles_ulcers
        medical_history.fainting_seizure = medical_history_data.fainting_seizure
        medical_history.rapid_weight_loss_radiation_therapy = medical_history_data.rapid_weight_loss_radiation_therapy
        medical_history.joint_replacement_implant = medical_history_data.joint_replacement_implant
        medical_history.heart_surgery_heart_attack = medical_history_data.heart_surgery_heart_attack
        medical_history.thyroid_problem = medical_history_data.thyroid_problem
        medical_history.heart_disease = medical_history_data.heart_disease
        medical_history.heart_murmur = medical_history_data.heart_murmur
        medical_history.hepatitis_liver_disease = medical_history_data.hepatitis_liver_disease
        medical_history.rheumatic_seizure = medical_history_data.rheumatic_seizure
        medical_history.respiratory_problems = medical_history_data.respiratory_problems
        medical_history.hepatitis_jaundice = medical_history_data.hepatitis_jaundice
        medical_history.tuberculosis = medical_history_data.tuberculosis
        medical_history.swollen_ankles = medical_history_data.swollen_ankles
        medical_history.kidney_disease = medical_history_data.kidney_disease
        medical_history.other_diseases = medical_history_data.other_diseases
        
        db.commit()
        db.refresh(medical_history)

        return {"message": "Medical history updated successfully", "medical_history": medical_history}
    
    
@app.post("/update_archive_status")
async def update_archive_status(archive_data: ArchiveUpdateRequest, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.id == archive_data.student_id).first()
    
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    student.is_archive = archive_data.is_archive

    if archive_data.is_archive:
        student.date_archived = date.today()
    else:
        student.date_archived = None

    db.commit()
    db.refresh(student)
    
    return {"message": "Student's archive status updated successfully", "status_code": 200}


@app.put("/update_active_status")
async def update_active_status(active_data: ActiveUpdateRequest, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.id == active_data.student_id).first()
    
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    student.is_active = active_data.is_active
    
    db.commit()
    
    return {"message": "Student's active status updated successfully"}


@app.post("/delete_student_record")
async def delete_student_record(delete: DeleteStudentRecord, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.id == delete.student_id).first()
    
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    db.query(MedicalHistory).filter(MedicalHistory.student_id == delete.student_id).delete()
    
    db.query(Appointment).filter(Appointment.student_id == delete.student_id).delete()

    db.query(OralHealthCondition).filter(OralHealthCondition.student_id == delete.student_id).delete()

    db.query(TemporaryTeeth).filter(TemporaryTeeth.student_id == delete.student_id).delete()

    db.query(PermanentTeeth).filter(PermanentTeeth.student_id == delete.student_id).delete()

    db.query(DentalProcedure).filter(DentalProcedure.student_id == delete.student_id).delete()

    db.query(ConditionTreatmentNeeds).filter(ConditionTreatmentNeeds.student_id == delete.student_id).delete()
    
    db.delete(student)
    db.commit()

    return {"message": "Student record deleted successfully", "status_code": 200}


@app.get("/download_parent_consent")
async def download_parent_consent():
    file_path = "static/PARENT CONSENT.pdf"

    return FileResponse(
        path=file_path,
        filename="PARENT CONSENT.pdf",
        media_type='application/pdf'
    )

def caesar_cipher_encrypt(text: str, shift: int) -> str:
    encrypted_text = []
    for char in text:
        if char.isalpha():
            shift_amount = shift % 26
            if char.islower():
                encrypted_text.append(chr((ord(char) - ord('a') + shift_amount) % 26 + ord('a')))
            elif char.isupper():
                encrypted_text.append(chr((ord(char) - ord('A') + shift_amount) % 26 + ord('A')))
        else:
            encrypted_text.append(char)
    return ''.join(encrypted_text)


def caesar_cipher_decrypt(encrypted_text: str, shift: int) -> str:
    decrypted_text = []
    for char in encrypted_text:
        if char.isalpha():
            shift_amount = shift % 26
            if char.islower():
                decrypted_text.append(chr((ord(char) - ord('a') - shift_amount) % 26 + ord('a')))
            elif char.isupper():
                decrypted_text.append(chr((ord(char) - ord('A') - shift_amount) % 26 + ord('A')))
        else:
            decrypted_text.append(char)
    return ''.join(decrypted_text)


@app.post('/forgot_password')
async def forgot_password(request: ForgotPasswordRequest, db: Session = Depends(get_database)):
    student = db.query(Student).filter(Student.email_address == request.email).first()
    
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    encrypted_email = caesar_cipher_encrypt(student.email_address, shift=3)

    # Generate the reset password URL with the encrypted email
    reset_url = f"https://ocnhs-dental.onrender.com/#/resetPassword/{encrypted_email}"
    # reset_url = f"http://localhost:5173/#/resetPassword/{encrypted_email}"
    
    msg = EmailMessage()
    msg['Subject'] = "Account Password Reset"
    msg['From'] = email_address
    msg['To'] = request.email
    msg.set_content(
    f"""\
    Hi {student.firstname} {student.lastname},

    We received a request to reset your password. If you made this request, just click the URL below to create a new password:

    URL: {reset_url}

    If you didn’t request a password reset, you can safely ignore this email — your account is still secure.

    This is an automated message—please do not reply.

    Thanks,
    OCNHS Dental Team
    """,

    )
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(email_address, email_password)
        smtp.send_message(msg)

    return { 'response': 'Forgot password successful.', 'status_code': 200 }


@app.post('/reset_password/{encrypted_email}')
async def reset_password(encrypted_email: str, request: ResetPasswordRequest, db: Session = Depends(get_database)):
    try:
        email = caesar_cipher_decrypt(encrypted_email, shift=3)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or corrupted link")

    student = db.query(Student).filter(Student.email_address == email).first()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    if request.new_password != request.confirm_new_password:
        raise HTTPException(status_code=400, detail="Password does not match")

    student.password = request.new_password
    db.commit()

    return { 'response': 'Reset password successful.', 'status_code': 200 }


@app.post("/create_temp_teeth")
async def create_temp_teeth(records: List[TemporaryTeethSchema],db: Session = Depends(get_database)):
    try:
        for record in records:
            existing = db.query(TemporaryTeeth).filter(TemporaryTeeth.student_id == record.student_id, TemporaryTeeth.grade_level == record.grade_level).first()

            if existing:
                existing.no_t_decayed = record.no_t_decayed
                existing.no_t_filled = record.no_t_filled
                existing.total_dft = record.total_dft
            else:
                new_record = TemporaryTeeth(
                    student_id=record.student_id,
                    grade_level=record.grade_level,
                    no_t_decayed=record.no_t_decayed,
                    no_t_filled=record.no_t_filled,
                    total_dft=record.total_dft
                )
                db.add(new_record)

        db.commit()

        return {"status_code": 200, "message": "Data saved successfully!"}
    except Exception as e:
        print("Error:", e)
        raise HTTPException(status_code=500, detail="Failed to save data.")


@app.get("/get_temporary_teeth/{student_id}")
def get_temporary_teeth(student_id: int, db: Session = Depends(get_database)):
    records = db.query(TemporaryTeeth).filter(TemporaryTeeth.student_id == student_id).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="Records not found")
    
    return records


@app.post("/create_perma_teeth")
async def create_perma_teeth(records: List[PermanentTeethSchema],db: Session = Depends(get_database)):
    try:
        for record in records:
            existing = db.query(PermanentTeeth).filter(PermanentTeeth.student_id == record.student_id, PermanentTeeth.grade_level == record.grade_level).first()

            if existing:
                existing.no_t_decayed = record.no_t_decayed
                existing.no_t_missing = record.no_t_missing
                existing.no_t_filled = record.no_t_filled
                existing.total_d_m_f_t = record.total_d_m_f_t
                existing.total_sound_teeth = record.total_sound_teeth
            else:
                new_record = PermanentTeeth(
                    student_id=record.student_id,
                    grade_level=record.grade_level,
                    no_t_decayed=record.no_t_decayed,
                    no_t_missing=record.no_t_missing,
                    no_t_filled=record.no_t_filled,
                    total_d_m_f_t=record.total_d_m_f_t,
                    total_sound_teeth=record.total_sound_teeth
                )
                db.add(new_record)

        db.commit()

        return {"status_code": 200, "message": "Data saved successfully!"}
    except Exception as e:
        print("Error:", e)
        raise HTTPException(status_code=500, detail="Failed to save data.")


@app.get("/get_permarnent_teeth/{student_id}")
def get_permarnent_teeth(student_id: int, db: Session = Depends(get_database)):
    records = db.query(PermanentTeeth).filter(PermanentTeeth.student_id == student_id).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="Records not found")
    
    return records


@app.post("/create_oral_health_condition")
async def create_oral_health_condition(records: List[OralHealthConditionSchema],db: Session = Depends(get_database)):
    try:
        for record in records:
            existing = db.query(OralHealthCondition).filter(OralHealthCondition.student_id == record.student_id, OralHealthCondition.grade_level == record.grade_level).first()

            if existing:
                existing.gingivitis = record.gingivitis
                existing.periodontal_disease = record.periodontal_disease
                existing.malocclusion = record.malocclusion
                existing.supernumerary_teeth = record.supernumerary_teeth
                existing.retained_deciduous_teeth = record.retained_deciduous_teeth
                existing.decubital_ulcer = record.decubital_ulcer
                existing.calculus = record.calculus
                existing.cleft_lip_palate = record.cleft_lip_palate
                existing.root_fragment = record.root_fragment
                existing.fluorosis = record.fluorosis
                existing.others = record.others
            else:
                new_record = OralHealthCondition(
                    student_id=record.student_id,
                    grade_level=record.grade_level,
                    gingivitis=record.gingivitis,
                    periodontal_disease=record.periodontal_disease,
                    malocclusion=record.malocclusion,
                    supernumerary_teeth=record.supernumerary_teeth,
                    retained_deciduous_teeth=record.retained_deciduous_teeth,
                    decubital_ulcer=record.decubital_ulcer,
                    calculus=record.calculus,
                    cleft_lip_palate=record.cleft_lip_palate,
                    root_fragment=record.root_fragment,
                    fluorosis=record.fluorosis,
                    others=record.others
                )
                db.add(new_record)

        db.commit()

        return {"status_code": 200, "message": "Data saved successfully!"}
    except Exception as e:
        print("Error:", e)
        raise HTTPException(status_code=500, detail="Failed to save data.")


@app.get("/get_oral_health_condition_teeth/{student_id}")
def get_oral_health_condition_teeth(student_id: int, db: Session = Depends(get_database)):
    records = db.query(OralHealthCondition).filter(OralHealthCondition.student_id == student_id).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="Records not found")
    
    return records


@app.post("/create_dental_procedure")
async def create_dental_procedure(records: List[DentalProcedureSchema],db: Session = Depends(get_database)):
    try:
        for record in records:
            existing = db.query(DentalProcedure).filter(DentalProcedure.student_id == record.student_id, DentalProcedure.grade_level == record.grade_level).first()

            if existing:
                existing.date = record.date
                existing.examination = record.examination
                existing.sealant_gi = record.sealant_gi
                existing.gum_treatment = record.gum_treatment
                existing.permanent_filling = record.permanent_filling
                existing.art = record.art
                existing.extraction = record.extraction
                existing.oral_prophylaxis = record.oral_prophylaxis
                existing.referral = record.referral
                existing.other_oral_treatment = record.other_oral_treatment
                existing.examined_by = record.examined_by
            else:
                new_record = DentalProcedure(
                    student_id=record.student_id,
                    grade_level=record.grade_level,
                    date=record.date,
                    examination=record.examination,
                    sealant_gi=record.sealant_gi,
                    gum_treatment=record.gum_treatment,
                    permanent_filling=record.permanent_filling,
                    art=record.art,
                    extraction=record.extraction,
                    oral_prophylaxis=record.oral_prophylaxis,
                    referral=record.referral,
                    other_oral_treatment=record.other_oral_treatment,
                    examined_by=record.examined_by
                )
                db.add(new_record)

        db.commit()

        return {"status_code": 200, "message": "Data saved successfully!"}
    except Exception as e:
        print("Error:", e)
        raise HTTPException(status_code=500, detail="Failed to save data.")


@app.get("/get_dental_procedure_teeth/{student_id}")
def get_dental_procedure_teeth(student_id: int, db: Session = Depends(get_database)):
    records = db.query(DentalProcedure).filter(DentalProcedure.student_id == student_id).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="Records not found")
    
    return records


@app.post("/create_condition_treatment_need")
async def create_condition_treatment_need(record: ConditionTreatmentNeedsSchema, db: Session = Depends(get_database)):
    existing = db.query(ConditionTreatmentNeeds).filter_by(student_id=record.student_id, layer_no=record.layer_no, cell_no=record.cell_no).first()

    if existing:
        existing.value = record.value
    else:
        new_record = ConditionTreatmentNeeds(
            student_id=record.student_id,
            layer_no=record.layer_no,
            cell_no=record.cell_no,
            value=record.value
        )
        db.add(new_record)

    db.commit()
    return {"status_code": 200, "message": "Data saved successfully!"}


@app.post("/create_condition_treatment_needs_array")
async def create_condition_treatment_needs_array(record: ConditionTreatmentNeedsArraySchema,db: Session = Depends(get_database)):
    try:
        db.query(ConditionTreatmentNeeds).filter_by(
            student_id=record.student_id,
            layer_no=record.layer_no,
            cell_no=record.cell_no
        ).delete()

        for value in record.values:
            db.add(ConditionTreatmentNeeds(
                student_id=record.student_id,
                layer_no=record.layer_no,
                cell_no=record.cell_no,
                value=value
            ))

        db.commit()
        return {"status_code": 200, "message": "Condition treatment needs updated successfully."}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/get_condition_treatment_needs/{student_id}")
async def get_condition_treatment_needs(student_id: int, db: Session = Depends(get_database)):
    records = db.query(ConditionTreatmentNeeds).filter_by(student_id=student_id).all()

    from collections import defaultdict
    value_map = defaultdict(list)
    for r in records:
        key = (r.layer_no, r.cell_no)
        value_map[key].append(r.value)

    result = {}

    for layer in range(1, 7):
        layer_cells = []
        for cell in range(1, 17):
            key = (layer, cell)
            values = value_map.get(key)

            if layer in [3, 4]:
                value = values if values else []
            else:
                value = values[0] if values else None

            layer_cells.append({
                "student_id": student_id,
                "layer_no": layer,
                "cell_no": cell,
                "value": value
            })

        result[f"layer_{layer}"] = layer_cells

    return result


@app.get('/generate_report')
async def generate_report(db: Session = Depends(get_database)):
    manila_tz = timezone("Asia/Manila")
    now = datetime.now(manila_tz)

    # Define types
    types = ['CLEANING', 'FLOURIDE', 'RESTORATION', 'EXTRACTION']

    # === TIME RANGES ===
    # Monthly
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_next_month = (start_of_month.replace(month=now.month + 1, day=1)
                           if now.month < 12 else start_of_month.replace(year=now.year + 1, month=1, day=1))

    # Quarterly
    quarter = (now.month - 1) // 3 + 1
    start_month = 3 * (quarter - 1) + 1
    start_of_quarter = now.replace(month=start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_next_quarter = (start_of_quarter.replace(month=start_month + 3, day=1)
                             if start_month + 3 <= 12 else start_of_quarter.replace(year=now.year + 1, month=(start_month + 3) % 12, day=1))

    # Yearly
    start_of_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_next_year = now.replace(year=now.year + 1, month=1, day=1)

    # === RESULTS CONTAINER ===
    report = {
        "monthly": {},
        "quarterly": {},
        "yearly": {}
    }

    # === QUERIES FOR EACH TYPE ===
    for appt_type in types:
        # Monthly
        monthly_count = db.query(Appointment).filter(
            Appointment.appointment_type == appt_type,
            Appointment.appointment_datetime >= start_of_month,
            Appointment.appointment_datetime < start_of_next_month
        ).count()

        # Quarterly
        quarterly_count = db.query(Appointment).filter(
            Appointment.appointment_type == appt_type,
            Appointment.appointment_datetime >= start_of_quarter,
            Appointment.appointment_datetime < start_of_next_quarter
        ).count()

        # Yearly
        yearly_count = db.query(Appointment).filter(
            Appointment.appointment_type == appt_type,
            Appointment.appointment_datetime >= start_of_year,
            Appointment.appointment_datetime < start_of_next_year
        ).count()

        # Store in report
        report["monthly"][appt_type] = monthly_count
        report["quarterly"][appt_type] = quarterly_count
        report["yearly"][appt_type] = yearly_count


    # Path to Excel file
    file_path = os.path.join(os.path.dirname(__file__), "static", "DENTAL ACCOMPLISHMENT REPORT.xlsx")

    # Load workbook and select active sheet
    wb = load_workbook(file_path)
    ws = wb.active  # Or specify sheet name with wb["SheetName"]

    # Define category order
    category_order = ['FLOURIDE', 'CLEANING', 'EXTRACTION', 'RESTORATION']

    # Row mapping
    row_mapping = {
        "monthly": 15,
        "quarterly": 16,
        "yearly": 17
    }

    # Column base
    base_column = 7  # G = 7

    # Insert data
    for period, row in row_mapping.items():
        total = 0
        # Put current year in column B
        ws.cell(row=row, column=2, value=now.year)

        for idx, category in enumerate(category_order):
            value = report[period][category]
            ws.cell(row=row, column=base_column + idx, value=value)
            total += value

        # Total in column K
        ws.cell(row=row, column=base_column + 4, value=total)

    # Save the Excel file
    wb.save(file_path)
    
    file_path = os.path.join(os.path.dirname(__file__), "static", "DENTAL ACCOMPLISHMENT REPORT.xlsx")
    return FileResponse(path=file_path, filename="DENTAL ACCOMPLISHMENT REPORT.xlsx", media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

